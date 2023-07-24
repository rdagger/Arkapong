"""Utility to convert levels designed in Excel to Python lists."""
import openpyxl
from os import path

# Replace file_path with the actual path to your Excel file
FILE_PATH = path.join("utils", "ArkaPong_Levels.xlsx")

# Define the mapping of colors to letters
COLOR_MAPPING = {
    'FFEEEEEE': 'W',
    'FFFF251F': 'R',
    'FF008EFF': 'B',
    'FFFB6F03': 'O',
    'FF0FE0F8': 'C',
    'FF11F50C': 'G',
    'FFF9EA00': 'Y',
    'FFFF3EF8': 'M',
    'FF9E9E9E': 'S',
    'FFBF9F48': 'A'
}

ENEMY_MAPPING = {
    "annulus": "0",
    "molecule": "1",
    "morph": "2",
    "orb": "3",
    "tetrahedron": "4"
}

GRID_HEIGHT = 10
GRID_WIDTH = 15


def read_excel_spreadsheet(filename, worksheet='Sheet1'):
    """Read Excel spreadsheet into a list.

        Args:
            filename(string): Spreadsheet filename
            worksheet(string): Worksheet name (default=Sheet1)
        """
    workbook = openpyxl.load_workbook(filename)
    # Select the desired worksheet
    sheet = workbook[worksheet]  # Replace 'Sheet1' with the name of your sheet
    data = []
    # Find the last populated row in the worksheet
    last_row = sheet.max_row
    start_row = 1
    for _ in range(1, last_row, GRID_HEIGHT):
        for row in range(start_row, start_row + GRID_HEIGHT - 1):
            row_data = []
            for col in range(1, GRID_WIDTH + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.comment:  # Enemy
                    letter = ENEMY_MAPPING[cell.comment.text]
                else:  # Brick
                    cell_color = (cell.fill.start_color.rgb
                                  if cell.fill.start_color.rgb else None)
                    letter = (COLOR_MAPPING.get(cell_color, '.')
                              if cell_color else '.')
                row_data.append(letter)
            data.append(row_data)
        # Skip a row between sets of columns
        data.append(['.'] * 15)
        start_row += 10
    workbook.close()
    return data


def get_column_breaks(grid):
    """Return list of column breaks between shapes.

    Args:
        grid([[string]]): Grid of level brick layout.
    Notes:
        Column breaks are defined by a column of all periods.
    """
    width = len(grid[0])
    empty_columns = []
    for col in range(width):
        if all(row[col] == '.' for row in grid):
            empty_columns.append(col)
    return empty_columns


def is_shape_blank(shape):
    """Determines if shape has no bricks.

    Args:
        shape([[string]]): List defining shape.
    """
    for sub_list in shape:
        for element in sub_list:
            if element != '.':
                return False
    return True


def split_shapes(grid, column_breaks):
    """Splits a level grid into individual shapes.

    Args:
        grid([[string]]): Grid of level brick layout.
        column_breaks([int]): List of the column breaks between shapes.
    """
    if not len(column_breaks):
        return [grid]
    shapes = []
    start_column = 0
    color_columns = 0
    column_count = len(grid[0])
    for column in range(column_count):
        if column in column_breaks or column == column_count-1:
            if color_columns > 0 or column == column_count-1:
                selected_columns = [row[start_column:column+1] for row in grid]
                shapes.append(selected_columns)
                start_column = column + 1
                color_columns = 0
        else:
            color_columns += 1
    # Remove any empty shapes (no elements)
    shapes = [shape for shape in shapes if shape]
    # Remove any blank shapes (all blank elements)
    for shape in shapes:
        if is_shape_blank(shape):
            shapes.remove(shape)
    return shapes


def trim_rows(shape):
    """Trim any blank rows from top and bottom of shape.

    Args:
        shape([[string]]): List defining shape.
    """
    trimmed_shape = []
    start_index = 0
    end_index = len(shape) - 1
    # Find the starting index of non-period rows
    for i, row in enumerate(shape):
        if any(elem != '.' for elem in row):
            start_index = i
            break
    # Find the ending index of non-period rows
    for i in range(len(shape) - 1, -1, -1):
        if any(elem != '.' for elem in shape[i]):
            end_index = i
            break
    # Create the trimmed shape
    trimmed_shape = shape[start_index: end_index + 1]
    return trimmed_shape


def trim_columns(shape):
    """Trim any blank columns from left and right of shape.

    Args:
        shape([[string]]): List defining shape.
    """
    trimmed_shape = []
    start_index = 0
    end_index = len(shape[0]) - 1
    # Find the starting index of non-period columns
    for col in range(len(shape[0])):
        if any(row[col] != '.' for row in shape):
            start_index = col
            break
    # Find the ending index of non-period columns
    for col in range(len(shape[0]) - 1, -1, -1):
        if any(row[col] != '.' for row in shape):
            end_index = col
            break
    # Create the trimmed shape by slicing the columns
    for row in shape:
        trimmed_row = row[start_index: end_index + 1]
        trimmed_shape.append(trimmed_row)
    return trimmed_shape


# Load level grids from spreadsheet
spreadsheet_data = read_excel_spreadsheet(FILE_PATH)
# Build levels
for i in range(0, len(spreadsheet_data), GRID_HEIGHT):
    grid = spreadsheet_data[i:i+GRID_HEIGHT]
    # Find shape column breaks
    column_breaks = get_column_breaks(grid)
    # Split shapes
    shapes = split_shapes(grid, column_breaks)
    # Trim blank rows and columns from each shape
    shapes = [trim_rows(trim_columns(shape)) for shape in shapes]
    # Print level definition
    print(f"LEVEL{i//GRID_HEIGHT + 1:03d} = [")
    last_index = len(shapes) - 1
    # Print all shape definitions in level
    for index, shape in enumerate(shapes):
        print("    [", end='')
        last_row = len(shape) - 1
        for irow, row in enumerate(shape):
            if irow > 0:
                print("     ", end='')
            print(f"{row}", end='')
            print("]", end='') if irow == last_row else print(",")
        print("]") if index == last_index else print(",")
    print()
